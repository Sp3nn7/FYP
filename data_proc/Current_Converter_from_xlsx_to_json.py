import json
import openpyxl


def converter(file_name):
    wb = openpyxl.load_workbook(file_name)
    all_data = []
    cities = []
    categories = []
    counter = 0
    regions = []
    unit_size = []
    for city in wb:
        print(city)
        current_city_data = []
        cities.append(str(city).split('"')[1])
        regions.append([city.cell(row=1, column=i).value for i in range(2, city.max_column+1)])
        for k in range(2, 8048):
            if counter == 0:
                categories.append(city.cell(row=k, column=1).value)
                unit_size.append(city.cell(row=k, column=1).value.split(" ")[-1])
            dummy = [city.cell(row=k, column=i).value for i in range(3, city.max_column+1)]
            data = []
            for x in dummy:
                if x != None:
                    data.append(x)
                else:
                    data.append(x)
            current_city_data.append(data)
        counter += 1
        all_data.append(current_city_data)

    dictionary = {
        "data":
            all_data,
        "datamap": {
            "city": cities,
            "category": categories,
            "unit": unit_size,
            "regions": regions
        }
    }

    json_object = json.dumps(dictionary, indent=4)
    with open("current_data.json", "w") as outfile:
        outfile.write(json_object)


converter('FoldedRegions.xlsx')
